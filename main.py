
from fastmcp import FastMCP
from openpyxl import load_workbook
from typing import TypedDict,Literal,List, Dict, Any,Optional
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart, Reference
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart.label import DataLabelList
from scipy import stats
import re
mcp=FastMCP(name="Excel MCP")


@mcp.tool
def create_sheet(url,sheet_name:str) -> None:
    """use to create a new sheet in an existing excel file"""
    wb=load_workbook(url)
    ws=wb.create_sheet(title=sheet_name)

    wb.save(url)

@mcp.tool
def get_all_sheets_name(url) ->list[str]:
    """use to get all sheet names in an existing excel file"""
    wb=load_workbook(url)
    return wb.sheetnames

@mcp.tool
def create_dashboard(
    title: str,
    x_axis_title: str,
    y_axis_title: str,
    data_sheet_name: str,
    categorical_column_number: int,
    value_column_number: int,
    position_of_chart: Literal["A5","J5","A20","J20"],
    url: str,
    chart_type: Literal["bar", "line", "pie", "area", "scatter"],
    title_of_dashboard: str

) -> None:
    """
    Create a chart in a 'Dashboard' sheet within an Excel workbook.
    
    Args:
        title: Chart title
        x_axis_title: X-axis label (ignored for pie charts)
        y_axis_title: Y-axis label (ignored for pie charts)
        data_sheet_name: Source sheet name containing data
        categorical_column_number: Column number for categories (1-based, e.g., 1 for column A)
        value_column_number: Column number for values (1-based)
        position_of_chart: Cell position for chart placement (e.g., "A1", "C5")
        url: Path to Excel file
        chart_type: Type of chart - "bar", "line", "pie", "area", or "scatter"
    """
    wb = load_workbook(url)
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard")

    dashboard = wb["Dashboard"]
    data_sheet = wb[data_sheet_name]

   
    cats = Reference(data_sheet, min_col=categorical_column_number, min_row=2, max_row=data_sheet.max_row)
    vals = Reference(data_sheet, min_col=value_column_number, min_row=1, max_row=data_sheet.max_row)


    if chart_type == "bar":
        chart = BarChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)  # Keep this for series names
        chart.set_categories(cats)
        chart.type = "col"
        chart.style = 4
        chart.shape = 4
        chart.y_axis.majorGridlines = None
        
        # Configure data labels to show ONLY values
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True        # Show value
        chart.dataLabels.showSerName = False   # Hide series name
        chart.dataLabels.showCatName = False   # Hide category name
        chart.dataLabels.showLegendKey = False # Hide legend key
    
        chart.x_axis.title.txPr = None
        chart.y_axis.title.txPr = None
        chart.title.txPr = None
    
    elif chart_type == "line":
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
    
    elif chart_type == "pie":
        chart = PieChart()
        chart.title = title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
    
    elif chart_type == "area":
        chart = AreaChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
    
    elif chart_type == "scatter":
        chart = ScatterChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)


    dashboard.merge_cells('A1:R4')
    
    cell = dashboard.cell(row=1, column=1)  
    cell.value =title_of_dashboard
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    cell.font  = Font(b=True, color="F8F8F8",size = 46)
    cell.fill = PatternFill("solid", fgColor="2591DB")

    dashboard.add_chart(chart, position_of_chart)
    chart.width = 15
    wb.save(url)
     


# 2 by 3 dashbaord code 
@mcp.tool
def create_dashboard_with_six_chart(
    title: str,
    x_axis_title: str,
    y_axis_title: str,
    data_sheet_name: str,
    categorical_column_number: int,
    value_column_number: int,
    position_of_chart: Literal["A5","I5","Q5","A20","I20","Q20"],
    url: str,
    chart_type: Literal["bar", "line", "pie", "area", "scatter"],
    title_of_dashboard: str

) -> None:
    """
    Create a chart in a 'Dashboard' sheet within an Excel workbook.
    
    Args:
        title: Chart title
        x_axis_title: X-axis label (ignored for pie charts)
        y_axis_title: Y-axis label (ignored for pie charts)
        data_sheet_name: Source sheet name containing data
        categorical_column_number: Column number for categories (1-based, e.g., 1 for column A)
        value_column_number: Column number for values (1-based)
        position_of_chart: Cell position for chart placement (e.g., "A1", "C5")
        url: Path to Excel file
        chart_type: Type of chart - "bar", "line", "pie", "area", or "scatter"
    """
    wb = load_workbook(url)
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard")

    dashboard = wb["Dashboard"]
    data_sheet = wb[data_sheet_name]

   
    cats = Reference(data_sheet, min_col=categorical_column_number, min_row=2, max_row=data_sheet.max_row)
    vals = Reference(data_sheet, min_col=value_column_number, min_row=1, max_row=data_sheet.max_row)


    if chart_type == "bar":
        chart = BarChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)  # Keep this for series names
        chart.set_categories(cats)
        chart.type = "col"
        chart.style = 4
        chart.shape = 4
        chart.y_axis.majorGridlines = None
        
        # Configure data labels to show ONLY values
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True        # Show value
        chart.dataLabels.showSerName = False   # Hide series name
        chart.dataLabels.showCatName = False   # Hide category name
        chart.dataLabels.showLegendKey = False # Hide legend key
    
        chart.x_axis.title.txPr = None
        chart.y_axis.title.txPr = None
        chart.title.txPr = None
    
    elif chart_type == "line":
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
    
    elif chart_type == "pie":
        chart = PieChart()
        chart.title = title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
    
    elif chart_type == "area":
        chart = AreaChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
    
    elif chart_type == "scatter":
        chart = ScatterChart()
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)


    dashboard.merge_cells('A1:R4')
    
    cell = dashboard.cell(row=1, column=1)  
    cell.value =title_of_dashboard
    cell.alignment = Alignment(horizontal='center', vertical='center')   
    cell.font  = Font(b=True, color="F8F8F8",size = 46)
    cell.fill = PatternFill("solid", fgColor="2591DB")

    dashboard.add_chart(chart, position_of_chart)
    chart.width = 13   # Width in inches (reduced from 15 to fit 3 charts)
    chart.height = 8
    wb.save(url)



@mcp.tool
def create_pivot(
    url: str,
    sheet_name: str,
    values_column: str,
    index_column: str,
    aggfunc: Literal["sum", "mean", "count", "min", "max", "median", "std"],
    pivot_sheet_name: str = "Pivot",
    sort_order: Literal["asc", "desc", "none"] = "none",
    top_k: Optional[int] = None
) -> None:
    """
    Create a pivot table from an existing Excel file and save it to a new sheet.
    
    Parameters:
    - url: Path to the Excel file
    - sheet_name: Name of the sheet containing the data
    - values_column: Column name for values to aggregate
    - index_column: Column name to use as pivot index
    - aggfunc: Aggregation function (sum, mean, count, min, max, median, std)
    - pivot_sheet_name: Name for the pivot sheet (default: "Pivot")
    - sort_order: Sort order for results - "asc" (ascending), "desc" (descending), or "none" (default: "none")
    - top_k: Return only top K results after sorting (optional, requires sort_order to be set)
    """
    
    df = pd.read_excel(url, sheet_name=sheet_name)
    pivot = df.pivot_table(values=values_column, index=index_column, aggfunc=aggfunc)
    
    # Apply sorting if specified
    if sort_order != "none":
        ascending = True if sort_order == "asc" else False
        pivot = pivot.sort_values(by=values_column, ascending=ascending)
    
    # Apply top K filter if specified
    if top_k is not None:
        if sort_order == "none":
            raise ValueError("sort_order must be 'asc' or 'desc' when using top_k parameter")
        pivot = pivot.head(top_k)
    
    wb = load_workbook(url)

    if pivot_sheet_name in wb.sheetnames:
        del wb[pivot_sheet_name]
    wb.create_sheet(pivot_sheet_name)
    
    ws = wb[pivot_sheet_name]
    
    ws.append([index_column, values_column])

    for index, value in pivot[values_column].items():
        ws.append([index, value])

    wb.save(url)

@mcp.tool
def create_histogram(
    url: str,
    data_sheet_name: str,
    value_column_number: int,
    num_bins: int,
    title: str,
    position_of_chart: Literal["A5","J5","A20","J20"],
    title_of_dashboard: str
) -> None:
    """Create histogram showing frequency distribution"""
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    col_name = df.columns[value_column_number - 1]
    
    hist_data = pd.cut(df[col_name], bins=num_bins).value_counts().sort_index()
    
    wb = load_workbook(url)
    
    hist_sheet_name = "HistogramData"
    if hist_sheet_name in wb.sheetnames:
        del wb[hist_sheet_name]
    wb.create_sheet(hist_sheet_name)
    
    hist_sheet = wb[hist_sheet_name]
    hist_sheet.append(["Range", "Frequency"])
    
    for interval, freq in hist_data.items():
        hist_sheet.append([str(interval), freq])
    
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard")
    
    dashboard = wb["Dashboard"]
    
    cats = Reference(hist_sheet, min_col=1, min_row=2, max_row=num_bins + 1)
    vals = Reference(hist_sheet, min_col=2, min_row=1, max_row=num_bins + 1)
    
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.x_axis.title = "Range"
    chart.y_axis.title = "Frequency"
    chart.style = 4
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    
    dashboard.merge_cells('A1:R4')
    cell = dashboard.cell(row=1, column=1)
    cell.value = title_of_dashboard
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(b=True, color="F8F8F8", size=46)
    cell.fill = PatternFill("solid", fgColor="2591DB")
    
    dashboard.add_chart(chart, position_of_chart)
    chart.width = 15
    
    wb.save(url)


@mcp.tool
def create_table(
    url: str,
    source_sheet_name: str,
    table_name: str,
    table_sheet_name: str = "TableSheet",
    start_cell: str = "A1",
    table_style: Literal[
        "TableStyleLight1", "TableStyleLight9", "TableStyleLight11",
        "TableStyleMedium2", "TableStyleMedium9", "TableStyleMedium15",
        "TableStyleDark1", "TableStyleDark10"
    ] = "TableStyleMedium9",
    show_row_stripes: bool = True,
    show_column_stripes: bool = False,
    show_first_column: bool = False,
    show_last_column: bool = False
) -> None:
    """
    Create a formatted table in a new sheet from existing data.
    
    Parameters:
    - url: Path to the Excel file
    - source_sheet_name: Name of the sheet containing source data
    - table_name: Name for the table (must be unique in workbook)
    - table_sheet_name: Name for the new sheet with table (default: "TableSheet")
    - start_cell: Starting cell for the table (default: "A1")
    - table_style: Table style name (default: "TableStyleMedium9")
    - show_row_stripes: Show alternating row colors (default: True)
    - show_column_stripes: Show alternating column colors (default: False)
    - show_first_column: Highlight first column (default: False)
    - show_last_column: Highlight last column (default: False)
    """
    
    # Load workbook
    wb = load_workbook(url)
    
    # Get source data
    source_sheet = wb[source_sheet_name]
    
    # Collect all data from source sheet
    data = []
    for row in source_sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # Skip empty rows
            data.append(list(row))
    
    # Create or clear table sheet
    if table_sheet_name in wb.sheetnames:
        del wb[table_sheet_name]
    wb.create_sheet(table_sheet_name)
    
    table_sheet = wb[table_sheet_name]
    
    # Write data to new sheet
    for row in data:
        table_sheet.append(row)
    
    # Calculate table range
    max_row = len(data)
    max_col = len(data[0]) if data else 1
    
    # Convert column number to letter
    def col_number_to_letter(n):
        result = ""
        while n > 0:
            n -= 1
            result = chr(n % 26 + 65) + result
            n //= 26
        return result
    
    end_col = col_number_to_letter(max_col)
    start_row = int(start_cell[1:]) if start_cell[1:].isdigit() else 1
    end_row = start_row + max_row - 1
    
    table_ref = f"{start_cell}:{end_col}{end_row}"
    
    # Create table
    table = Table(displayName=table_name, ref=table_ref)
    
    # Apply style
    style = TableStyleInfo(
        name=table_style,
        showFirstColumn=show_first_column,
        showLastColumn=show_last_column,
        showRowStripes=show_row_stripes,
        showColumnStripes=show_column_stripes
    )
    table.tableStyleInfo = style
    
    # Add table to sheet
    table_sheet.add_table(table)
    
    # Save workbook
    wb.save(url)

@mcp.tool
def get_sheet_preview(url: str, sheet_name: str, num_rows: int = 10) -> Dict[str, Any]:
    """
    Get top N rows of an Excel sheet with metadata.
    This provides enough context for an LLM to understand the data structure
    and create dashboards without exceeding token limits.
    
    Args:
        url: Path to Excel file
        sheet_name: Name of the sheet
        num_rows: Number of data rows to return (default 10, excluding header)
    
    Returns:
        Dictionary with preview data and basic metadata
    """
    wb = load_workbook(url, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    
    # Get total dimensions
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    
    # Extract headers (row 1)
    headers = []
    for col in range(1, total_cols + 1):
        cell_value = sheet.cell(1, col).value
        headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")
    
    # Extract top N data rows (rows 2 to num_rows+1)
    rows = []
    for row_idx in range(2, min(num_rows + 2, total_rows + 1)):
        row_data = []
        for col in range(1, total_cols + 1):
            cell_value = sheet.cell(row_idx, col).value
            row_data.append(cell_value)
        rows.append(row_data)
    
    wb.close()
    
    return {
        "sheet_name": sheet_name,
        "total_rows": total_rows - 1,  # Excluding header
        "total_columns": total_cols,
        "headers": headers,
        "preview_rows": rows,
        "showing_rows": len(rows)
    }

@mcp.tool
def create_combo_chart(
    url: str,
    data_sheet_name: str,
    title: str,
    x_axis_title: str,
    y_axis_title: str,
    secondary_y_axis_title: str,
    categorical_column_number: int,
    bar_value_column_number: int,
    line_value_column_number: int,
    position_of_chart: Literal["A5","J5","A20","J20"],
    title_of_dashboard: str
) -> None:
    """Create a combination chart with bar and line series"""
    wb = load_workbook(url)
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard")
    
    dashboard = wb["Dashboard"]
    data_sheet = wb[data_sheet_name]
    
    cats = Reference(data_sheet, min_col=categorical_column_number, min_row=2, max_row=data_sheet.max_row)
    bar_vals = Reference(data_sheet, min_col=bar_value_column_number, min_row=1, max_row=data_sheet.max_row)
    line_vals = Reference(data_sheet, min_col=line_value_column_number, min_row=1, max_row=data_sheet.max_row)
    
    chart = BarChart()
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.add_data(bar_vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.type = "col"
    chart.style = 4
    
    line_chart = LineChart()
    line_chart.y_axis.title = secondary_y_axis_title
    line_chart.add_data(line_vals, titles_from_data=True)
    line_chart.set_categories(cats)
    
    chart += line_chart
    chart.y_axis.axId = 100
    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"
    
    dashboard.merge_cells('A1:R4')
    cell = dashboard.cell(row=1, column=1)
    cell.value = title_of_dashboard
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(b=True, color="F8F8F8", size=46)
    cell.fill = PatternFill("solid", fgColor="2591DB")
    
    dashboard.add_chart(chart, position_of_chart)
    chart.width = 15
    wb.save(url)

@mcp.tool
def create_stacked_chart(
    url: str,
    data_sheet_name: str,
    title: str,
    x_axis_title: str,
    y_axis_title: str,
    categorical_column_number: int,
    value_column_numbers: List[int],
    position_of_chart: Literal["A5","J5","A20","J20"],
    chart_type: Literal["bar", "area"],
    title_of_dashboard: str
) -> None:
    """Create stacked bar or area chart for multiple series"""
    wb = load_workbook(url)
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard")
    
    dashboard = wb["Dashboard"]
    data_sheet = wb[data_sheet_name]
    
    cats = Reference(data_sheet, min_col=categorical_column_number, min_row=2, max_row=data_sheet.max_row)
    
    if chart_type == "bar":
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
    else:
        chart = AreaChart()
        chart.grouping = "stacked"
    
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.style = 4
    
    for col_num in value_column_numbers:
        vals = Reference(data_sheet, min_col=col_num, min_row=1, max_row=data_sheet.max_row)
        chart.add_data(vals, titles_from_data=True)
    
    chart.set_categories(cats)
    
    dashboard.merge_cells('A1:R4')
    cell = dashboard.cell(row=1, column=1)
    cell.value = title_of_dashboard
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(b=True, color="F8F8F8", size=46)
    cell.fill = PatternFill("solid", fgColor="2591DB")
    
    dashboard.add_chart(chart, position_of_chart)
    chart.width = 15
    wb.save(url)


@mcp.tool
def drop_duplicates(
    url:str,
    data_sheet_name:str,
) -> None:
    """
    Remove duplicate rows from an Excel sheet and save the cleaned data back to the same file.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    """
    df=pd.read_excel(url,sheet_name=data_sheet_name)
    df.drop_duplicates(inplace=True)
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def drop_nulls(
    url:str,
    data_sheet_name:str,
) -> None:
    """
    Remove rows with null values from an Excel sheet and save the cleaned data back to the same file.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    """
    df=pd.read_excel(url,sheet_name=data_sheet_name)
    df.dropna(inplace=True)
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def fill_nulls(
        url:str,
        data_sheet_name:str,
        column_to_fill:str,
        fill_value:Literal["mean","median","mode","zero"],
        fill_coustome_value:str = None
) -> None:
    """
    Fill null values in a specified column of an Excel sheet and save the updated data back to the same file.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_to_fill: Column name where null values need to be filled
    - fill_value: Method to fill nulls ("mean", "median", "mode", "zero")
    - fill_coustome_value: Custom value to fill nulls if specified
    """
    df=pd.read_excel(url,sheet_name=data_sheet_name)
    
    if fill_value == "mean":
        fill_val = df[column_to_fill].mean()
    elif fill_value == "median":
        fill_val = df[column_to_fill].median()
    elif fill_value == "mode":
        fill_val = df[column_to_fill].mode()[0]
    elif fill_value == "zero":
        fill_val = 0
    else:
        fill_val = fill_coustome_value

    df[column_to_fill].fillna(fill_val, inplace=True)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def dataset_summary_for_data_claning(url: str, data_sheet_name: str) -> dict:

    """
    Generate a summary of the dataset in the specified Excel sheet for data cleaning purposes.
    Parameters:
    - url: Path to the Excel file  
    - data_sheet_name: Name of the sheet containing the data
    Returns:
    - A dictionary containing summary statistics and information about the dataset."""
   
    df=pd.read_excel(url,sheet_name=data_sheet_name)

    summary = {}

    # Basic shape
    summary["rows"] = df.shape[0]
    summary["columns"] = df.shape[1]

    # Column details
    summary["column_info"] = []
    for col in df.columns:
        summary["column_info"].append({
            "name": col,
            "dtype": str(df[col].dtype),
            "null_values": int(df[col].isnull().sum()),
            "unique_values": int(df[col].nunique()),
            "sample_values": df[col].dropna().unique()[:5].tolist()
        })

    # Statistics for numeric columns
    summary["statistics"] = df.describe(include='all').to_dict()

    # Duplicate count
    summary["duplicate_rows"] = int(df.duplicated().sum())

    # Memory usage
    summary["memory_usage_kb"] = round(df.memory_usage(deep=True).sum() / 1024, 2)

    return summary

@mcp.tool
def trim_whitespace(
    url: str,
    data_sheet_name: str,
    columns: Optional[List[str]] = None
) -> None:
    """
    Remove leading and trailing whitespace from specified columns.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - columns: List of column names to trim (None = all text columns)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if columns is None:
        columns = df.select_dtypes(include=['object']).columns.tolist()
    
    for col in columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def standardize_case(
    url: str,
    data_sheet_name: str,
    column_name: str,
    case_type: Literal["lower", "upper", "title"]
) -> None:
    """
    Standardize text case in a column.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to standardize
    - case_type: Type of case conversion (lower, upper, title)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if case_type == "lower":
        df[column_name] = df[column_name].astype(str).str.lower()
    elif case_type == "upper":
        df[column_name] = df[column_name].astype(str).str.upper()
    elif case_type == "title":
        df[column_name] = df[column_name].astype(str).str.title()
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def remove_special_characters(
    url: str,
    data_sheet_name: str,
    column_name: str,
    keep_spaces: bool = True
) -> None:
    """
    Remove special characters from a text column.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to clean
    - keep_spaces: Whether to preserve spaces (default: True)
    """
    import re
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if keep_spaces:
        df[column_name] = df[column_name].astype(str).apply(
            lambda x: re.sub(r'[^a-zA-Z0-9\s]', '', x)
        )
    else:
        df[column_name] = df[column_name].astype(str).apply(
            lambda x: re.sub(r'[^a-zA-Z0-9]', '', x)
        )
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def standardize_text_values(
    url: str,
    data_sheet_name: str,
    column_name: str,
    mapping_dict: str
) -> None:
    """
    Standardize text values using a mapping dictionary.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to standardize
    - mapping_dict: JSON string of mappings e.g., '{"yes": "Yes", "y": "Yes"}'
    """
    import json
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    mapping = json.loads(mapping_dict)
    
    df[column_name] = df[column_name].replace(mapping)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def convert_data_types(
    url: str,
    data_sheet_name: str,
    column_name: str,
    target_type: Literal["int", "float", "string", "bool"]
) -> None:
    """
    Convert column data type.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to convert
    - target_type: Target data type (int, float, string, bool)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if target_type == "int":
        df[column_name] = pd.to_numeric(df[column_name], errors='coerce').astype('Int64')
    elif target_type == "float":
        df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
    elif target_type == "string":
        df[column_name] = df[column_name].astype(str)
    elif target_type == "bool":
        df[column_name] = df[column_name].astype(bool)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def extract_numeric(
    url: str,
    data_sheet_name: str,
    column_name: str,
    new_column_name: str
) -> None:
    """
    Extract numeric values from text columns.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to extract from
    - new_column_name: Name for the new numeric column
    """
    import re
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    df[new_column_name] = df[column_name].astype(str).apply(
        lambda x: float(re.sub(r'[^\d.]', '', x)) if re.search(r'\d', x) else None
    )
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def detect_outliers(
    url: str,
    data_sheet_name: str,
    column_name: str,
    method: Literal["iqr", "zscore", "percentile"],
    threshold: float = 1.5
) -> Dict[str, Any]:
    """
    Detect outliers in a numeric column.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to check for outliers
    - method: Detection method (iqr, zscore, percentile)
    - threshold: Threshold value (1.5 for IQR, 3 for Z-score)
    
    Returns:
    - Dictionary with outlier information
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    outliers = []
    
    if method == "iqr":
        Q1 = df[column_name].quantile(0.25)
        Q3 = df[column_name].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - threshold * IQR
        upper_bound = Q3 + threshold * IQR
        outliers = df[(df[column_name] < lower_bound) | (df[column_name] > upper_bound)].index.tolist()
        
    elif method == "zscore":
        
        z_scores = stats.zscore(df[column_name].dropna())
        outliers = df[column_name].dropna()[abs(z_scores) > threshold].index.tolist()
        
    elif method == "percentile":
        lower = df[column_name].quantile(threshold / 100)
        upper = df[column_name].quantile(1 - threshold / 100)
        outliers = df[(df[column_name] < lower) | (df[column_name] > upper)].index.tolist()
    
    return {
        "column": column_name,
        "method": method,
        "outlier_count": len(outliers),
        "outlier_indices": outliers[:100]  # Limit to first 100
    }

@mcp.tool
def handle_outliers(
    url: str,
    data_sheet_name: str,
    column_name: str,
    method: Literal["remove", "cap", "replace"],
    detection_method: Literal["iqr", "zscore"] = "iqr",
    threshold: float = 1.5
) -> None:
    """
    Handle outliers by removing, capping, or replacing them.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to handle outliers
    - method: How to handle outliers (remove, cap, replace)
    - detection_method: How to detect outliers (iqr, zscore)
    - threshold: Detection threshold
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if detection_method == "iqr":
        Q1 = df[column_name].quantile(0.25)
        Q3 = df[column_name].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - threshold * IQR
        upper_bound = Q3 + threshold * IQR
    else:  # zscore
        #from scipy import stats
        z_scores = abs(stats.zscore(df[column_name].dropna()))
        mean = df[column_name].mean()
        std = df[column_name].std()
        lower_bound = mean - threshold * std
        upper_bound = mean + threshold * std
    
    if method == "remove":
        df = df[(df[column_name] >= lower_bound) & (df[column_name] <= upper_bound)]
    elif method == "cap":
        df[column_name] = df[column_name].clip(lower=lower_bound, upper=upper_bound)
    elif method == "replace":
        median = df[column_name].median()
        df.loc[(df[column_name] < lower_bound) | (df[column_name] > upper_bound), column_name] = median
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def rename_columns(
    url: str,
    data_sheet_name: str,
    column_mapping: str
) -> None:
    """
    Rename columns using a mapping dictionary.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_mapping: JSON string e.g., '{"old_name": "new_name"}'
    """
    import json
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    mapping = json.loads(column_mapping)
    
    df.rename(columns=mapping, inplace=True)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def drop_columns(
    url: str,
    data_sheet_name: str,
    columns_to_drop: List[str]
) -> None:
    """
    Drop specified columns from the sheet.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - columns_to_drop: List of column names to drop
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    df.drop(columns=columns_to_drop, inplace=True, errors='ignore')
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def reorder_columns(
    url: str,
    data_sheet_name: str,
    column_order: List[str]
) -> None:
    """
    Reorder columns in the specified order.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_order: List of column names in desired order
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    # Add any columns not in the order list at the end
    remaining_cols = [col for col in df.columns if col not in column_order]
    final_order = column_order + remaining_cols
    
    df = df[final_order]
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def split_column(
    url: str,
    data_sheet_name: str,
    column_name: str,
    delimiter: str,
    new_column_names: List[str]
) -> None:
    """
    Split a column into multiple columns.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to split
    - delimiter: Character to split on
    - new_column_names: List of names for new columns
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    split_data = df[column_name].astype(str).str.split(delimiter, expand=True)
    
    for i, new_col in enumerate(new_column_names):
        if i < len(split_data.columns):
            df[new_col] = split_data[i]
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def merge_columns(
    url: str,
    data_sheet_name: str,
    columns_to_merge: List[str],
    new_column_name: str,
    separator: str = " "
) -> None:
    """
    Merge multiple columns into one.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - columns_to_merge: List of column names to merge
    - new_column_name: Name for the new merged column
    - separator: Character to join with (default: space)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    df[new_column_name] = df[columns_to_merge].astype(str).agg(separator.join, axis=1)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def replace_values(
    url: str,
    data_sheet_name: str,
    column_name: str,
    old_value: str,
    new_value: str
) -> None:
    """
    Replace specific values in a column.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to modify
    - old_value: Value to replace
    - new_value: Replacement value
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    df[column_name] = df[column_name].replace(old_value, new_value)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def map_values(
    url: str,
    data_sheet_name: str,
    column_name: str,
    value_mapping: str
) -> None:
    """
    Map values using a dictionary.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to map
    - value_mapping: JSON string e.g., '{"old1": "new1", "old2": "new2"}'
    """
    import json
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    mapping = json.loads(value_mapping)
    
    df[column_name] = df[column_name].map(mapping).fillna(df[column_name])
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def standardize_categories(
    url: str,
    data_sheet_name: str,
    column_name: str,
    category_mapping: str
) -> None:
    """
    Standardize categorical values by grouping similar categories.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name with categories
    - category_mapping: JSON string e.g., '{"cat1": "Category1", "cat 1": "Category1"}'
    """
    import json
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    mapping = json.loads(category_mapping)
    
    df[column_name] = df[column_name].replace(mapping)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def fill_forward_backward(
    url: str,
    data_sheet_name: str,
    column_name: str,
    method: Literal["ffill", "bfill"]
) -> None:
    """
    Fill missing values using forward or backward fill.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to fill
    - method: Fill method (ffill = forward fill, bfill = backward fill)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if method == "ffill":
        df[column_name] = df[column_name].ffill()
    elif method == "bfill":
        df[column_name] = df[column_name].bfill()
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def interpolate_missing(
    url: str,
    data_sheet_name: str,
    column_name: str,
    method: Literal["linear", "polynomial"] = "linear"
) -> None:
    """
    Interpolate missing numeric values.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to interpolate
    - method: Interpolation method (linear, polynomial)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if method == "linear":
        df[column_name] = df[column_name].interpolate(method='linear')
    elif method == "polynomial":
        df[column_name] = df[column_name].interpolate(method='polynomial', order=2)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def validate_range(
    url: str,
    data_sheet_name: str,
    column_name: str,
    min_value: float,
    max_value: float
) -> Dict[str, Any]:
    """
    Validate if numeric values are within expected range.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to validate
    - min_value: Minimum acceptable value
    - max_value: Maximum acceptable value
    
    Returns:
    - Dictionary with validation results
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    out_of_range = df[(df[column_name] < min_value) | (df[column_name] > max_value)]
    
    return {
        "column": column_name,
        "min_value": min_value,
        "max_value": max_value,
        "total_rows": len(df),
        "out_of_range_count": len(out_of_range),
        "out_of_range_indices": out_of_range.index.tolist()[:100]
    }

@mcp.tool
def validate_format(
    url: str,
    data_sheet_name: str,
    column_name: str,
    pattern_type: Literal["email", "phone", "url", "custom"],
    custom_pattern: Optional[str] = None
) -> Dict[str, Any]:
    """
    Validate if text matches expected format.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to validate
    - pattern_type: Type of pattern (email, phone, url, custom)
    - custom_pattern: Regex pattern if pattern_type is custom
    
    Returns:
    - Dictionary with validation results
    """
    import re
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    patterns = {
        "email": r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$',
        "phone": r'^\+?1?\d{9,15}$',
        "url": r'^https?://[^\s/$.?#].[^\s]*$'
    }
    
    pattern = custom_pattern if pattern_type == "custom" else patterns.get(pattern_type)
    
    invalid = df[~df[column_name].astype(str).str.match(pattern, na=False)]
    
    return {
        "column": column_name,
        "pattern_type": pattern_type,
        "total_rows": len(df),
        "invalid_count": len(invalid),
        "invalid_indices": invalid.index.tolist()[:100]
    }

@mcp.tool
def find_inconsistent_values(
    url: str,
    data_sheet_name: str,
    column_name: str
) -> Dict[str, Any]:
    """
    Find inconsistent values in categorical columns like case variations and spacing issues.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to check
    
    Returns:
    - Dictionary with unique values, counts, and similar groups
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    value_counts = df[column_name].value_counts().to_dict()
    unique_values = df[column_name].unique().tolist()
    
    # Find similar values (case-insensitive or with extra spaces)
    similar_groups = {}
    processed = set()
    
    for val in unique_values:
        if val not in processed and isinstance(val, str):
            normalized = val.strip().lower()
            similar = [v for v in unique_values if isinstance(v, str) and v.strip().lower() == normalized]
            if len(similar) > 1:
                similar_groups[val] = similar
                processed.update(similar)
    
    return {
        "column": column_name,
        "unique_count": len(unique_values),
        "value_counts": value_counts,
        "similar_groups": similar_groups
    }


@mcp.tool
def drop_duplicates_by_columns(
    url: str,
    data_sheet_name: str,
    columns: List[str],
    keep: Literal["first", "last", "none"] = "first"
) -> None:
    """
    Drop duplicates based on specific columns only.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - columns: List of column names to check for duplicates
    - keep: Which duplicate to keep (first, last, none = remove all duplicates)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    keep_value = False if keep == "none" else keep
    df.drop_duplicates(subset=columns, keep=keep_value, inplace=True)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def mark_duplicates(
    url: str,
    data_sheet_name: str,
    columns: Optional[List[str]] = None,
    flag_column_name: str = "is_duplicate"
) -> None:
    """
    Mark duplicate rows with a boolean flag column instead of removing them.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - columns: List of columns to check (None = all columns)
    - flag_column_name: Name for the flag column (default: "is_duplicate")
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if columns:
        df[flag_column_name] = df.duplicated(subset=columns, keep=False)
    else:
        df[flag_column_name] = df.duplicated(keep=False)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def filter_rows_by_condition(
    url: str,
    data_sheet_name: str,
    column_name: str,
    operator: Literal["==", "!=", ">", "<", ">=", "<=", "contains", "not_contains"],
    value: str,
    keep: bool = True
) -> None:
    """
    Filter rows based on a condition.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to filter on
    - operator: Comparison operator (==, !=, >, <, >=, <=, contains, not_contains)
    - value: Value to compare against
    - keep: True to keep matching rows, False to remove matching rows
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    if operator == "==":
        mask = df[column_name] == value
    elif operator == "!=":
        mask = df[column_name] != value
    elif operator == ">":
        mask = df[column_name] > float(value)
    elif operator == "<":
        mask = df[column_name] < float(value)
    elif operator == ">=":
        mask = df[column_name] >= float(value)
    elif operator == "<=":
        mask = df[column_name] <= float(value)
    elif operator == "contains":
        mask = df[column_name].astype(str).str.contains(value, na=False)
    elif operator == "not_contains":
        mask = ~df[column_name].astype(str).str.contains(value, na=False)
    
    df = df[mask] if keep else df[~mask]
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def remove_empty_rows(
    url: str,
    data_sheet_name: str
) -> None:
    """
    Remove rows where all values are null/empty.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    df.dropna(how='all', inplace=True)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def normalize_numeric(
    url: str,
    data_sheet_name: str,
    column_name: str,
    method: Literal["minmax", "zscore"],
    new_column_name: Optional[str] = None
) -> None:
    """
    Normalize numeric data using min-max or z-score scaling.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to normalize
    - method: Normalization method (minmax: scales to 0-1, zscore: standardizes with mean=0, std=1)
    - new_column_name: Name for normalized column (None = overwrite original)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    target_col = new_column_name if new_column_name else column_name
    
    if method == "minmax":
        min_val = df[column_name].min()
        max_val = df[column_name].max()
        df[target_col] = (df[column_name] - min_val) / (max_val - min_val)
    elif method == "zscore":
        mean = df[column_name].mean()
        std = df[column_name].std()
        df[target_col] = (df[column_name] - mean) / std
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def round_numeric(
    url: str,
    data_sheet_name: str,
    column_name: str,
    decimals: int = 2
) -> None:
    """
    Round numeric values to specified decimal places.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name to round
    - decimals: Number of decimal places (default: 2)
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    df[column_name] = df[column_name].round(decimals)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def standardize_phone_numbers(
    url: str,
    data_sheet_name: str,
    column_name: str,
    format_type: Literal["international", "national", "digits_only"]
) -> None:
    """
    Standardize phone number formats.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name with phone numbers
    - format_type: Format type (international: +1-XXX-XXX-XXXX, national: (XXX) XXX-XXXX, digits_only: XXXXXXXXXX)
    """
    import re
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    def format_phone(phone):
        if pd.isna(phone):
            return phone
        
        # Extract digits only
        digits = re.sub(r'\D', '', str(phone))
        
        if format_type == "digits_only":
            return digits
        elif format_type == "international" and len(digits) == 10:
            return f"+1-{digits[:3]}-{digits[3:6]}-{digits[6:]}"
        elif format_type == "national" and len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        else:
            return phone
    
    df[column_name] = df[column_name].apply(format_phone)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def standardize_emails(
    url: str,
    data_sheet_name: str,
    column_name: str,
    remove_invalid: bool = False
) -> None:
    """
    Standardize and validate email addresses.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name with emails
    - remove_invalid: Whether to remove rows with invalid emails (default: False)
    """
    import re
    
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'


    df[column_name] = df[column_name].astype(str).str.lower().str.strip()
    
    if remove_invalid:
        df = df[df[column_name].str.match(email_pattern, na=False)]
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def standardize_date_format(
    url: str,
    data_sheet_name: str,
    column_name: str,
    target_format: Literal["YYYY-MM-DD", "DD/MM/YYYY", "MM/DD/YYYY", "DD-MM-YYYY", "MM-DD-YYYY", "YYYY/MM/DD"],
    error_handling: Literal["coerce", "raise", "ignore"] = "coerce"
) -> None:
    """
    Convert date columns to a consistent, standardized format.
    This tool handles various input date formats and converts them to your desired output format.
    Useful for ensuring consistency across date columns and preparing data for analysis.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - column_name: Column name with dates to standardize
    - target_format: Desired output date format
        * "YYYY-MM-DD" → 2024-12-17 (ISO format, best for sorting)
        * "DD/MM/YYYY" → 17/12/2024 (European format)
        * "MM/DD/YYYY" → 12/17/2024 (US format)
        * "DD-MM-YYYY" → 17-12-2024
        * "MM-DD-YYYY" → 12-17-2024
        * "YYYY/MM/DD" → 2024/12/17
    - error_handling: How to handle invalid dates
        * "coerce" → Convert invalid dates to NaT (Not a Time)
        * "raise" → Raise error if invalid date found
        * "ignore" → Keep original value if conversion fails
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    # Convert to datetime with error handling
    df[column_name] = pd.to_datetime(df[column_name], errors=error_handling)
    
    # Map target format to strftime format
    format_map = {
        "YYYY-MM-DD": "%Y-%m-%d",
        "DD/MM/YYYY": "%d/%m/%Y",
        "MM/DD/YYYY": "%m/%d/%Y",
        "DD-MM-YYYY": "%d-%m-%Y",
        "MM-DD-YYYY": "%m-%d-%Y",
        "YYYY/MM/DD": "%Y/%m/%d"
    }
    
    # Apply the target format
    strftime_format = format_map[target_format]
    df[column_name] = df[column_name].dt.strftime(strftime_format)
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def parse_dates(
    url: str,
    data_sheet_name: str,
    date_column: str,
    components: List[Literal["year", "month", "day", "quarter", "week", "dayofweek", "dayofyear", "weekday_name", "month_name"]],
    prefix: str = ""
) -> None:
    """
    Extract date components (year, month, day, etc.) from a date column into separate columns.
    This is useful for time-based analysis, grouping, and creating features for analysis.
    Each component will be added as a new column.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - date_column: Column name containing dates to parse
    - components: List of date components to extract:
        * "year" → Extract year (e.g., 2024)
        * "month" → Extract month number (1-12)
        * "day" → Extract day of month (1-31)
        * "quarter" → Extract quarter (1-4)
        * "week" → Extract week of year (1-53)
        * "dayofweek" → Extract day of week as number (0=Monday, 6=Sunday)
        * "dayofyear" → Extract day of year (1-366)
        * "weekday_name" → Extract day name (Monday, Tuesday, etc.)
        * "month_name" → Extract month name (January, February, etc.)
    - prefix: Optional prefix for new column names (e.g., "date_" creates "date_year", "date_month")
    
    Example:
        If date_column = "OrderDate" with value "2024-12-17"
        and components = ["year", "month", "day"]
        Creates: OrderDate_year=2024, OrderDate_month=12, OrderDate_day=17
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    # Convert to datetime if not already
    df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
    
    # Extract requested components
    for component in components:
        col_name = f"{prefix}{date_column}_{component}" if prefix else f"{date_column}_{component}"
        
        if component == "year":
            df[col_name] = df[date_column].dt.year
        elif component == "month":
            df[col_name] = df[date_column].dt.month
        elif component == "day":
            df[col_name] = df[date_column].dt.day
        elif component == "quarter":
            df[col_name] = df[date_column].dt.quarter
        elif component == "week":
            df[col_name] = df[date_column].dt.isocalendar().week
        elif component == "dayofweek":
            df[col_name] = df[date_column].dt.dayofweek
        elif component == "dayofyear":
            df[col_name] = df[date_column].dt.dayofyear
        elif component == "weekday_name":
            df[col_name] = df[date_column].dt.day_name()
        elif component == "month_name":
            df[col_name] = df[date_column].dt.month_name()
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)


@mcp.tool
def fill_date_gaps(
    url: str,
    data_sheet_name: str,
    date_column: str,
    frequency: Literal["D", "W", "M", "Q", "Y", "H"] = "D",
    fill_method: Literal["ffill", "bfill", "interpolate", "zero", "none"] = "ffill",
    columns_to_fill: Optional[List[str]] = None
) -> None:
    """
    Identify and fill missing dates in time series data.
    This tool creates a complete date range and fills in any gaps with specified values.
    Essential for time series analysis where continuous date sequences are required.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - date_column: Column name containing dates
    - frequency: Time frequency for filling gaps:
        * "D" → Daily (fills missing days)
        * "W" → Weekly (fills missing weeks)
        * "M" → Monthly (fills missing months)
        * "Q" → Quarterly (fills missing quarters)
        * "Y" → Yearly (fills missing years)
        * "H" → Hourly (fills missing hours)
    - fill_method: How to fill values in other columns for missing dates:
        * "ffill" → Forward fill (use previous value)
        * "bfill" → Backward fill (use next value)
        * "interpolate" → Linear interpolation between values
        * "zero" → Fill with zeros
        * "none" → Leave as null/NaN
    - columns_to_fill: List of column names to fill (None = fill all numeric columns)
    
    Example:
        If you have sales data with dates: 2024-01-01, 2024-01-03, 2024-01-05
        This will add missing dates: 2024-01-02, 2024-01-04
        And fill their values based on fill_method
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    # Convert date column to datetime
    df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
    
    # Sort by date
    df = df.sort_values(by=date_column)
    
    # Create complete date range
    date_range = pd.date_range(
        start=df[date_column].min(),
        end=df[date_column].max(),
        freq=frequency
    )
    
    # Create a DataFrame with complete date range
    complete_dates = pd.DataFrame({date_column: date_range})
    
    # Merge with original data to identify gaps
    df = complete_dates.merge(df, on=date_column, how='left')
    
    # Determine which columns to fill
    if columns_to_fill is None:
        columns_to_fill = df.select_dtypes(include=['number']).columns.tolist()
    
    # Apply fill method to specified columns
    for col in columns_to_fill:
        if col in df.columns and col != date_column:
            if fill_method == "ffill":
                df[col] = df[col].ffill()
            elif fill_method == "bfill":
                df[col] = df[col].bfill()
            elif fill_method == "interpolate":
                df[col] = df[col].interpolate(method='linear')
            elif fill_method == "zero":
                df[col] = df[col].fillna(0)
            # "none" means leave as NaN, so no action needed
    
    with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

@mcp.tool
def execute_custom_pandas_code(
    url: str,
    data_sheet_name: str,
    custom_code: str,
    output_sheet_name: Optional[str] = None,
    return_preview: bool = True,
    preview_rows: int = 10
) -> Dict[str, Any]:
    """
    Execute custom pandas code on Excel data when standard tools don't cover your use case.
    This is a powerful fallback tool that gives you full pandas flexibility for complex data manipulations.
    
    ⚠️ IMPORTANT SAFETY NOTES:
    - This tool executes arbitrary Python code - use with caution
    - Only use when existing tools cannot handle your specific scenario
    - The code runs in a restricted namespace with pandas, numpy, and datetime available
    - Your code should work with the 'df' variable (DataFrame) which is automatically loaded
    - Your code must assign the result back to 'df' variable
    
    AVAILABLE VARIABLES IN YOUR CODE:
    - df: pandas DataFrame (automatically loaded from the Excel sheet)
    - pd: pandas library
    - np: numpy library
    - datetime: datetime module
    - re: regular expressions module
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    - custom_code: Your custom pandas code as a string
        * Must work with the 'df' variable
        * Must assign final result to 'df'
        * Can use multi-line code
    - output_sheet_name: Name for output sheet (None = overwrite original sheet)
    - return_preview: Whether to return a preview of the result (default: True)
    - preview_rows: Number of rows to return in preview (default: 10)
    
    Returns:
    - Dictionary with execution status, preview data, and any error messages
    
    EXAMPLES OF CUSTOM CODE:
    
    Example 1 - Complex date calculation:
    '''
    df['days_since_start'] = (df['end_date'] - df['start_date']).dt.days
    df['is_overdue'] = df['days_since_start'] > 30
    '''
    
    Example 2 - Advanced text processing:
    '''
    df['email_domain'] = df['email'].str.split('@').str[1]
    df['is_corporate'] = ~df['email_domain'].str.contains('gmail|yahoo|hotmail')
    '''
    
    Example 3 - Complex aggregation:
    '''
    df['running_total'] = df.groupby('category')['amount'].cumsum()
    df['pct_of_category'] = df['amount'] / df.groupby('category')['amount'].transform('sum') * 100
    '''
    
    Example 4 - Conditional logic:
    '''
    def categorize_age(age):
        if age < 18:
            return 'Minor'
        elif age < 65:
            return 'Adult'
        else:
            return 'Senior'
    
    df['age_group'] = df['age'].apply(categorize_age)
    '''
    
    Example 5 - Multiple operations:
    '''
    # Remove outliers using IQR method
    Q1 = df['sales'].quantile(0.25)
    Q3 = df['sales'].quantile(0.75)
    IQR = Q3 - Q1
    df = df[(df['sales'] >= Q1 - 1.5 * IQR) & (df['sales'] <= Q3 + 1.5 * IQR)]
    
    # Create new calculated column
    df['profit_margin'] = (df['revenue'] - df['cost']) / df['revenue'] * 100
    '''
    """
    import numpy as np
    from datetime import datetime
    
    try:
        # Load the data
        df = pd.read_excel(url, sheet_name=data_sheet_name)
        original_shape = df.shape
        
        # Create a restricted namespace for code execution
        namespace = {
            'df': df,
            'pd': pd,
            'np': np,
            'datetime': datetime,
            're': re,
            '__builtins__': {
                'len': len,
                'str': str,
                'int': int,
                'float': float,
                'bool': bool,
                'list': list,
                'dict': dict,
                'tuple': tuple,
                'set': set,
                'range': range,
                'enumerate': enumerate,
                'zip': zip,
                'map': map,
                'filter': filter,
                'sum': sum,
                'max': max,
                'min': min,
                'abs': abs,
                'round': round,
                'sorted': sorted,
                'any': any,
                'all': all,
                'print': print,
            }
        }
        
        # Execute the custom code
        exec(custom_code, namespace)
        
        # Get the modified DataFrame
        df = namespace['df']
        new_shape = df.shape
        
        # Determine output sheet name
        target_sheet = output_sheet_name if output_sheet_name else data_sheet_name
        
        # Save the result
        with pd.ExcelWriter(url, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=target_sheet, index=False)
        
        # Prepare response
        response = {
            "status": "success",
            "message": "Custom code executed successfully",
            "original_shape": {"rows": original_shape[0], "columns": original_shape[1]},
            "new_shape": {"rows": new_shape[0], "columns": new_shape[1]},
            "output_sheet": target_sheet,
            "columns": df.columns.tolist()
        }
        
        # Add preview if requested
        if return_preview:
            preview_data = df.head(preview_rows).to_dict(orient='records')
            response["preview"] = preview_data
            response["preview_rows"] = len(preview_data)
        
        return response
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"Error executing custom code: {str(e)}",
            "error_type": type(e).__name__,
            "suggestion": "Check your code syntax and ensure 'df' variable is used correctly"
        }


@mcp.tool
def validate_custom_code(
    custom_code: str
) -> Dict[str, Any]:
    """
    Validate custom pandas code before execution to catch syntax errors.
    This is a safety tool to check your code without actually running it on your data.
    Use this to test your code logic before using execute_custom_pandas_code.
    
    Parameters:
    - custom_code: Your custom pandas code as a string to validate
    
    Returns:
    - Dictionary with validation status and any syntax errors found
    
    Example usage:
    First validate: validate_custom_code("df['new_col'] = df['old_col'] * 2")
    Then execute: execute_custom_pandas_code(..., custom_code="df['new_col'] = df['old_col'] * 2")
    """
    import ast
    
    try:
        # Try to parse the code as an Abstract Syntax Tree
        ast.parse(custom_code)
        
        # Check for dangerous operations
        dangerous_keywords = ['import', 'open', 'eval', 'exec', '__', 'os.', 'sys.', 'subprocess']
        found_dangerous = [kw for kw in dangerous_keywords if kw in custom_code]
        
        if found_dangerous:
            return {
                "status": "warning",
                "message": "Code contains potentially dangerous operations",
                "dangerous_keywords": found_dangerous,
                "suggestion": "Avoid using imports, file operations, or system commands"
            }
        
        # Check if 'df' is used in the code
        if 'df' not in custom_code:
            return {
                "status": "warning",
                "message": "Code does not reference 'df' variable",
                "suggestion": "Make sure to use 'df' to work with your DataFrame"
            }
        
        return {
            "status": "success",
            "message": "Code syntax is valid and appears safe to execute",
            "code_lines": len(custom_code.split('\n'))
        }
        
    except SyntaxError as e:
        return {
            "status": "error",
            "message": f"Syntax error in code: {str(e)}",
            "line": e.lineno,
            "suggestion": "Fix the syntax error before executing"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"Validation error: {str(e)}",
            "error_type": type(e).__name__
        }


@mcp.tool
def get_dataframe_info(
    url: str,
    data_sheet_name: str
) -> Dict[str, Any]:
    """
    Get detailed information about the DataFrame structure to help write custom code.
    This tool helps you understand your data before writing custom pandas operations.
    Use this to see column names, data types, and sample values.
    
    Parameters:
    - url: Path to the Excel file
    - data_sheet_name: Name of the sheet containing the data
    
    Returns:
    - Dictionary with DataFrame metadata, columns info, and sample data
    """
    df = pd.read_excel(url, sheet_name=data_sheet_name)
    
    # Collect column information
    column_details = []
    for col in df.columns:
        column_details.append({
            "name": col,
            "dtype": str(df[col].dtype),
            "non_null_count": int(df[col].count()),
            "null_count": int(df[col].isnull().sum()),
            "unique_values": int(df[col].nunique()),
            "sample_values": df[col].dropna().head(3).tolist()
        })
    
    return {
        "shape": {"rows": df.shape[0], "columns": df.shape[1]},
        "columns": df.columns.tolist(),
        "column_details": column_details,
        "memory_usage_mb": round(df.memory_usage(deep=True).sum() / (1024 * 1024), 2),
        "sample_data": df.head(5).to_dict(orient='records')
    }




if __name__ == "__main__":
    mcp.run()